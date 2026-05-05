from openpyxl import Workbook, load_workbook2
import os

# Members:
# ALVAREZ, MARK JOSEPH A. - Menu, Add Student, Validation
# BASILA, ASHLEY D. - View, Search, Delete

FILE_NAME = "students.xlsx"

# Lists to store student information
names = []
ages = []


# Load data from Excel
def load_students():
    if os.path.exists(FILE_NAME):
        wb = load_workbook(FILE_NAME)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is not None and row[1] is not None:
                names.append(row[0])
                ages.append(row[1])


# Save data to Excel
def save_students():
    wb = Workbook()
    ws = wb.active
    ws.title = "Students"

    # Header
    ws.append(["Name", "Age"])

    # Student data
    for i in range(len(names)):
        ws.append([names[i], ages[i]])

    wb.save(FILE_NAME)


# Function to show menu
def menu():
    print("\n=== STUDENT INFORMATION SYSTEM ===")
    print("1. Add Student")
    print("2. View Students")
    print("3. Search Student")
    print("4. Delete Student")
    print("5. Exit")


# Function to add a student
def add_student():
    name = input("Enter student name: ").strip()
    if name == "":
        print("Name cannot be empty.")
        return

    age_input = input("Enter student age: ")
    if not age_input.isdigit():
        print("Age must be a number.")
        return

    age = int(age_input)
    if age <= 0:
        print("Age must be more than 0.")
        return

    names.append(name)
    ages.append(age)

    save_students()  # ito yung auto save
    print(f"Student '{name}' is added.")


# Ashley's functions
def view_students():
    if len(names) == 0:
        print("No students to display.")
    else:
        print("\n--- STUDENT LIST ---")
        for i in range(len(names)):
            print(f"{i+1}. Name: {names[i]}, Age: {ages[i]}")


def search_student():
    search_name = input("Enter name to search: ").strip()
    found = False

    for i in range(len(names)):
        if names[i].lower() == search_name.lower():
            print(f"Found: {names[i]}, Age: {ages[i]}")
            found = True
            break

    if not found:
        print("Student not found.")


def delete_student():
    del_name = input("Enter name to delete: ").strip()
    found = False

    for i in range(len(names)):
        if names[i].lower() == del_name.lower():
            confirm = input(f"Are you sure you want to delete {names[i]}? (y/n): ")

            if confirm.lower() == "y":
                names.pop(i)
                ages.pop(i)

                save_students()  # auto save pa din
                print("Student deleted successfully.")
            else:
                print("Delete cancelled.")

            found = True
            break

    if not found:
        print("Student not found.")


# Load saved students before starting
load_students()


# Main program loop
while True:
    menu()
    choice = input("Choose an option (1-5) only :> : ")

    if choice == "1":
        add_student()

    elif choice == "2":
        view_students()

    elif choice == "3":
        search_student()

    elif choice == "4":
        delete_student()

    elif choice == "5":
        save_students()  # save before exit naka auto save na din
        print("Exiting program. Bye!")
        break

    else:
        print("Wrong choice. Enter number 1 to 5.")
